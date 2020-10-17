import openpyxl

# gives us a new workbook with a single sheet
# wb = openpyxl.Workbook()

# to load an existing workbook
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)  # returns `['Sheet1']`

sheet = wb["Sheet1"]

# to create a new sheet
# wb.create_sheet("Sheet2", 0)

# remove a sheet
# wb.remove_sheet(sheet)

# accessing cells in a sheet
# e.g. first row, first object
# cell = sheet["a1"]
# print(cell.value)  # returns `transaction_id`

# change the value of the cell
# cell.value = 1

# print(cell.row)  # returns 1
# print(cell.column)  # returns 1
# print(cell.coordinate)  # returns A1

# instead of passing coordinate to the sheet object
cell = sheet.cell(row=1, column=1)
# print(sheet.max_row)  # returns 4
# print(sheet.max_column)  # returns 3

# we're starting range from 1, since we don't have an 0 row in spreadsheets
# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)

# using the square bracket syntax sheet["a1"]

# returns all the cells in the a column
column = sheet["a"]

# work with a range of cells
cells = sheet["a:c"]  # returns a tuple of cells

# we can also use coodrniates in our sheet
# sheet["a1:c3"]
sheet[1:3]

print(cells)

# used to add a row at the bottom of the sheet
# you pass a list or a tuple of values
sheet.append([1, 2, 3])

# insert a row at a given index
# sheet.insert_rows  # insert a row at a given index
# sheet.insert_cols
# sheet.delete_rows
# sheet.delete_cols

# save the workbook
wb.save("transactions2.xlsx")
