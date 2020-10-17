import openpyxl

wb = openpyxl.load_workbook("transactions.xlsx")

# this is an example of a command function
# as a result of calling this, the state of our system in this case
# our workbook changes
# so every time we call it we get a new sheet in our workbook
# wb.create_sheet()

sheet = wb["Sheet1"]

# we use this to access a cell in the sheet
# this is an example of a query method
# we use it to access a given cell
# however this violates the command query separation principle
# sheet.cell()

for row in range(1, 10):
    # when we try to access a row that doesn't exist it creates it for us
    # this is the violation of command query separation principle
    # because this is a query method,
    # it shouldn't change the state of our system in this case the workbook
    # in other words, it shouldn't have a side effect
    # the violation of this principle
    # can cause unexpected behaviours in your program
    # it will be better if this method raised an exception
    # when you try to access a cell that didn't exist,
    # similar to accessing the 4th item in a list of 3 items
    # so when working with the `openpyxl` package be aware of this
    # in case when you see some blank rows in your spreadsheets
    cell = sheet.cell(row, 1)
    print(cell.value)

sheet.append([1, 2, 3])

wb.save("transactions2.xlsx")
